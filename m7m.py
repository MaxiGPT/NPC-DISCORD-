import discord
from discord.ext import commands
from discord import app_commands
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import os
import asyncio
from datetime import datetime
from typing import Optional, List, Dict, Any
from dotenv import load_dotenv
import logging

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger('NPCBot')

# Cargar variables de entorno
load_dotenv()
DISCORD_TOKEN = os.getenv("DISCORD_TOKEN")

# Configuración del bot
intents = discord.Intents.default()
intents.message_content = True
intents.guilds = True

class ExcelManager:
    """Gestor para operaciones con archivos Excel"""
    
    def __init__(self, filename: str):
        self.filename = filename
        self.ensure_file_exists()
    
    def ensure_file_exists(self):
        """Asegura que el archivo Excel existe con la estructura correcta"""
        if not os.path.exists("data"):
            os.makedirs("data")
        
        filepath = os.path.join("data", self.filename)
        if not os.path.exists(filepath):
            wb = openpyxl.Workbook()
            ws = wb.active
            
            if "npcs" in self.filename:
                ws.title = "NPCs"
                headers = ["id", "nombre", "descripcion", "imagen_url", "rol", 
                          "canal_id", "inventario", "creador_id", "fecha_creacion", "mensaje_id"]
            else:  # items
                ws.title = "Items"
                headers = ["id", "nombre", "descripcion", "imagen_url", "categoria",
                          "propiedades", "creador_id", "fecha_creacion"]
            
            ws.append(headers)
            wb.save(filepath)
            logger.info(f"Archivo {filepath} creado con éxito")
    
    def get_workbook(self) -> openpyxl.Workbook:
        """Obtiene el workbook del archivo Excel"""
        filepath = os.path.join("data", self.filename)
        return openpyxl.load_workbook(filepath)
    
    def save_workbook(self, wb: openpyxl.Workbook):
        """Guarda el workbook"""
        filepath = os.path.join("data", self.filename)
        wb.save(filepath)
    
    def get_next_id(self) -> int:
        """Obtiene el siguiente ID disponible"""
        wb = self.get_workbook()
        ws = wb.active
        
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        
        return max_id + 1
    
    def get_all_records(self) -> List[Dict[str, Any]]:
        """Obtiene todos los registros como lista de diccionarios"""
        wb = self.get_workbook()
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        records = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Si tiene ID
                record = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        record[header] = row[i]
                records.append(record)
        
        return records
    
    def get_record_by_id(self, record_id: int) -> Optional[Dict[str, Any]]:
        """Obtiene un registro por ID"""
        records = self.get_all_records()
        for record in records:
            if record.get("id") == record_id:
                return record
        return None
    
    def add_record(self, data: Dict[str, Any]) -> int:
        """Añade un nuevo registro y retorna su ID"""
        wb = self.get_workbook()
        ws = wb.active
        
        new_id = self.get_next_id()
        data["id"] = new_id
        data["fecha_creacion"] = datetime.now().isoformat()
        
        headers = [cell.value for cell in ws[1]]
        row_data = [data.get(header, "") for header in headers]
        
        ws.append(row_data)
        self.save_workbook(wb)
        
        return new_id
    
    def update_record(self, record_id: int, updates: Dict[str, Any]) -> bool:
        """Actualiza un registro existente"""
        wb = self.get_workbook()
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == record_id:
                for header, value in updates.items():
                    if header in headers:
                        col_idx = headers.index(header) + 1
                        ws.cell(row=row_num, column=col_idx, value=value)
                
                self.save_workbook(wb)
                return True
        
        return False
    
    def delete_record(self, record_id: int) -> bool:
        """Elimina un registro"""
        wb = self.get_workbook()
        ws = wb.active
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == record_id:
                ws.delete_rows(row_num)
                self.save_workbook(wb)
                return True
        
        return False

class NPCModal(discord.ui.Modal):
    """Modal para crear/editar NPCs"""
    
    def __init__(self, title: str = "Crear NPC", npc_data: Optional[Dict] = None):
        super().__init__(title=title)
        
        self.nombre = discord.ui.TextInput(
            label="Nombre",
            placeholder="Nombre del NPC",
            default=npc_data.get("nombre", "") if npc_data else "",
            max_length=100
        )
        self.add_item(self.nombre)
        
        self.descripcion = discord.ui.TextInput(
            label="Descripción",
            style=discord.TextStyle.paragraph,
            placeholder="Descripción detallada del NPC",
            default=npc_data.get("descripcion", "") if npc_data else "",
            max_length=1000
        )
        self.add_item(self.descripcion)
        
        self.imagen_url = discord.ui.TextInput(
            label="URL de Imagen",
            placeholder="https://ejemplo.com/imagen.png",
            default=npc_data.get("imagen_url", "") if npc_data else "",
            required=False
        )
        self.add_item(self.imagen_url)

class ItemModal(discord.ui.Modal):
    """Modal para crear/editar ítems"""
    
    def __init__(self, title: str = "Crear Ítem", item_data: Optional[Dict] = None):
        super().__init__(title=title)
        
        self.nombre = discord.ui.TextInput(
            label="Nombre",
            placeholder="Nombre del ítem",
            default=item_data.get("nombre", "") if item_data else "",
            max_length=100
        )
        self.add_item(self.nombre)
        
        self.descripcion = discord.ui.TextInput(
            label="Descripción",
            style=discord.TextStyle.paragraph,
            placeholder="Descripción del ítem",
            default=item_data.get("descripcion", "") if item_data else "",
            max_length=1000
        )
        self.add_item(self.descripcion)
        
        self.imagen_url = discord.ui.TextInput(
            label="URL de Imagen",
            placeholder="https://ejemplo.com/imagen.png",
            default=item_data.get("imagen_url", "") if item_data else "",
            required=False
        )
        self.add_item(self.imagen_url)
        
        self.propiedades = discord.ui.TextInput(
            label="Propiedades",
            placeholder="ataque:5;defensa:2;velocidad:1",
            default=item_data.get("propiedades", "") if item_data else "",
            required=False
        )
        self.add_item(self.propiedades)

class RolSelect(discord.ui.Select):
    """Select para elegir rol del NPC"""
    
    def __init__(self, default_value: Optional[str] = None):
        options = [
            discord.SelectOption(label="Aliado", value="Aliado", 
                               default=default_value == "Aliado"),
            discord.SelectOption(label="Enemigo", value="Enemigo",
                               default=default_value == "Enemigo"),
            discord.SelectOption(label="Neutro", value="Neutro",
                               default=default_value == "Neutro")
        ]
        super().__init__(placeholder="Selecciona el rol del NPC", options=options)

class CategoriaSelect(discord.ui.Select):
    """Select para elegir categoría del ítem"""
    
    def __init__(self, default_value: Optional[str] = None):
        options = [
            discord.SelectOption(label="Arma", value="Arma",
                               default=default_value == "Arma"),
            discord.SelectOption(label="Consumible", value="Consumible",
                               default=default_value == "Consumible"),
            discord.SelectOption(label="Objeto Mágico", value="Objeto Mágico",
                               default=default_value == "Objeto Mágico"),
            discord.SelectOption(label="Accesorio", value="Accesorio",
                               default=default_value == "Accesorio")
        ]
        super().__init__(placeholder="Selecciona la categoría", options=options)

class PaginationView(discord.ui.View):
    """Vista para paginación de listas"""
    
    def __init__(self, embeds: List[discord.Embed]):
        super().__init__(timeout=180)
        self.embeds = embeds
        self.current_page = 0
        self.update_buttons()
    
    def update_buttons(self):
        self.previous_button.disabled = self.current_page == 0
        self.next_button.disabled = self.current_page >= len(self.embeds) - 1
    
    @discord.ui.button(label="◀️", style=discord.ButtonStyle.secondary)
    async def previous_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_page = max(0, self.current_page - 1)
        self.update_buttons()
        await interaction.response.edit_message(embed=self.embeds[self.current_page], view=self)
    
    @discord.ui.button(label="▶️", style=discord.ButtonStyle.secondary)
    async def next_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_page = min(len(self.embeds) - 1, self.current_page + 1)
        self.update_buttons()
        await interaction.response.edit_message(embed=self.embeds[self.current_page], view=self)

class NPCBot(commands.Bot):
    def __init__(self):
        super().__init__(command_prefix="/", intents=intents)
        self.npc_manager = ExcelManager("npcs.xlsx")
        self.item_manager = ExcelManager("items.xlsx")
    
    async def setup_hook(self):
        await self.add_cog(NPCCommands(self))
        await self.add_cog(ItemCommands(self))
        logger.info("Cogs cargados correctamente")

class NPCCommands(commands.Cog):
    def __init__(self, bot: NPCBot):
        self.bot = bot
    
    @app_commands.command(name="npc", description="Gestión de NPCs")
    @app_commands.describe(accion="Acción a realizar")
    async def npc_command(self, interaction: discord.Interaction, accion: str):
        """Comando principal para NPCs con subcomandos"""
        await interaction.response.send_message(
            "Usa uno de los siguientes comandos:\n"
            "`/npc_crear`, `/npc_listar`, `/npc_ver`, `/npc_editar`, `/npc_eliminar`",
            ephemeral=True
        )
    
    @app_commands.command(name="npc_crear", description="Crea un nuevo NPC")
    async def npc_crear(self, interaction: discord.Interaction):
        """Crear un nuevo NPC"""
        modal = NPCModal(title="Crear Nuevo NPC")
        
        async def on_submit(modal_interaction: discord.Interaction):
            await modal_interaction.response.defer(ephemeral=True)
            
            # Crear vista para seleccionar rol y canal
            view = discord.ui.View()
            rol_select = RolSelect()
            
            # Select de canales
            channels = [ch for ch in interaction.guild.text_channels][:25]
            channel_options = [
                discord.SelectOption(label=ch.name, value=str(ch.id))
                for ch in channels
            ]
            channel_select = discord.ui.Select(
                placeholder="Selecciona el canal",
                options=channel_options
            )
            
            view.add_item(rol_select)
            view.add_item(channel_select)
            
            embed = discord.Embed(
                title="Configuración del NPC",
                description="Selecciona el rol y el canal para el NPC",
                color=discord.Color.blue()
            )
            
            await modal_interaction.followup.send(embed=embed, view=view, ephemeral=True)
            
            # Esperar selección
            def check(i):
                return i.user == interaction.user
            
            try:
                select_interaction = await self.bot.wait_for("interaction", check=check, timeout=60)
                await select_interaction.response.defer()
                
                # Obtener valores seleccionados
                rol = rol_select.values[0] if rol_select.values else "Neutro"
                canal_id = channel_select.values[0] if channel_select.values else None
                
                # Crear NPC
                npc_data = {
                    "nombre": modal.nombre.value,
                    "descripcion": modal.descripcion.value,
                    "imagen_url": modal.imagen_url.value or "",
                    "rol": rol,
                    "canal_id": canal_id,
                    "inventario": "",
                    "creador_id": str(interaction.user.id)
                }
                
                npc_id = self.bot.npc_manager.add_record(npc_data)
                
                # Crear embed de confirmación
                confirm_embed = discord.Embed(
                    title="✅ NPC Creado Exitosamente",
                    description=f"**{npc_data['nombre']}** (ID: {npc_id})",
                    color=discord.Color.green()
                )
                confirm_embed.add_field(name="Rol", value=rol, inline=True)
                confirm_embed.add_field(name="Canal", value=f"<#{canal_id}>", inline=True)
                if npc_data["imagen_url"]:
                    confirm_embed.set_thumbnail(url=npc_data["imagen_url"])
                
                await select_interaction.followup.send(embed=confirm_embed)
                
                # Crear ficha en el canal
                if canal_id:
                    channel = self.bot.get_channel(int(canal_id))
                    if channel:
                        ficha_embed = self.crear_embed_npc(npc_data, npc_id)
                        msg = await channel.send(embed=ficha_embed)
                        
                        # Actualizar mensaje_id
                        self.bot.npc_manager.update_record(npc_id, {"mensaje_id": str(msg.id)})
                
            except asyncio.TimeoutError:
                await modal_interaction.followup.send("⏱️ Tiempo agotado", ephemeral=True)
        
        modal.on_submit = on_submit
        await interaction.response.send_modal(modal)
    
    @app_commands.command(name="npc_listar", description="Lista todos los NPCs")
    @app_commands.describe(rol="Filtrar por rol (opcional)")
    async def npc_listar(self, interaction: discord.Interaction, 
                        rol: Optional[str] = None):
        """Listar NPCs con paginación"""
        await interaction.response.defer()
        
        npcs = self.bot.npc_manager.get_all_records()
        
        # Filtrar por rol si se especifica
        if rol:
            npcs = [npc for npc in npcs if npc.get("rol") == rol]
        
        if not npcs:
            await interaction.followup.send("❌ No se encontraron NPCs")
            return
        
        # Crear embeds paginados
        embeds = []
        items_per_page = 5
        
        for i in range(0, len(npcs), items_per_page):
            embed = discord.Embed(
                title=f"📋 Lista de NPCs" + (f" - Rol: {rol}" if rol else ""),
                color=discord.Color.blue()
            )
            
            page_npcs = npcs[i:i + items_per_page]
            for npc in page_npcs:
                canal_text = f"<#{npc.get('canal_id')}>" if npc.get('canal_id') else "Sin canal"
                creador_text = f"<@{npc.get('creador_id')}>" if npc.get('creador_id') else "Desconocido"
                
                embed.add_field(
                    name=f"{npc['nombre']} (ID: {npc['id']})",
                    value=f"**Rol:** {npc.get('rol', 'N/A')}\n"
                          f"**Canal:** {canal_text}\n"
                          f"**Creador:** {creador_text}",
                    inline=False
                )
            
            embed.set_footer(text=f"Página {len(embeds) + 1} de {((len(npcs) - 1) // items_per_page) + 1}")
            embeds.append(embed)
        
        if len(embeds) == 1:
            await interaction.followup.send(embed=embeds[0])
        else:
            view = PaginationView(embeds)
            await interaction.followup.send(embed=embeds[0], view=view)
    
    @app_commands.command(name="npc_ver", description="Ver detalles de un NPC")
    @app_commands.describe(npc_id="ID del NPC")
    async def npc_ver(self, interaction: discord.Interaction, npc_id: int):
        """Ver información detallada de un NPC"""
        await interaction.response.defer()
        
        npc = self.bot.npc_manager.get_record_by_id(npc_id)
        if not npc:
            await interaction.followup.send("❌ NPC no encontrado")
            return
        
        embed = self.crear_embed_npc(npc, npc_id)
        await interaction.followup.send(embed=embed)
    
    @app_commands.command(name="npc_editar", description="Editar un NPC existente")
    @app_commands.describe(npc_id="ID del NPC a editar")
    async def npc_editar(self, interaction: discord.Interaction, npc_id: int):
        """Editar un NPC"""
        npc = self.bot.npc_manager.get_record_by_id(npc_id)
        if not npc:
            await interaction.response.send_message("❌ NPC no encontrado", ephemeral=True)
            return
        
        # Verificar permisos
        if str(interaction.user.id) != npc.get("creador_id") and not interaction.user.guild_permissions.manage_channels:
            await interaction.response.send_message("❌ No tienes permisos para editar este NPC", ephemeral=True)
            return
        
        modal = NPCModal(title=f"Editar NPC: {npc['nombre']}", npc_data=npc)
        
        async def on_submit(modal_interaction: discord.Interaction):
            await modal_interaction.response.defer(ephemeral=True)
            
            updates = {
                "nombre": modal.nombre.value,
                "descripcion": modal.descripcion.value,
                "imagen_url": modal.imagen_url.value or ""
            }
            
            if self.bot.npc_manager.update_record(npc_id, updates):
                embed = discord.Embed(
                    title="✅ NPC Actualizado",
                    description=f"**{updates['nombre']}** ha sido actualizado",
                    color=discord.Color.green()
                )
                await modal_interaction.followup.send(embed=embed)
                
                # Actualizar ficha en el canal si existe
                if npc.get("mensaje_id") and npc.get("canal_id"):
                    try:
                        channel = self.bot.get_channel(int(npc["canal_id"]))
                        if channel:
                            msg = await channel.fetch_message(int(npc["mensaje_id"]))
                            updated_npc = self.bot.npc_manager.get_record_by_id(npc_id)
                            new_embed = self.crear_embed_npc(updated_npc, npc_id)
                            await msg.edit(embed=new_embed)
                    except:
                        pass
            else:
                await modal_interaction.followup.send("❌ Error al actualizar el NPC", ephemeral=True)
        
        modal.on_submit = on_submit
        await interaction.response.send_modal(modal)
    
    @app_commands.command(name="npc_eliminar", description="Eliminar un NPC")
    @app_commands.describe(npc_id="ID del NPC a eliminar")
    async def npc_eliminar(self, interaction: discord.Interaction, npc_id: int):
        """Eliminar un NPC"""
        await interaction.response.defer(ephemeral=True)
        
        npc = self.bot.npc_manager.get_record_by_id(npc_id)
        if not npc:
            await interaction.followup.send("❌ NPC no encontrado")
            return
        
        # Verificar permisos
        if str(interaction.user.id) != npc.get("creador_id") and not interaction.user.guild_permissions.manage_channels:
            await interaction.followup.send("❌ No tienes permisos para eliminar este NPC")
            return
        
        # Confirmación
        embed = discord.Embed(
            title="⚠️ Confirmar Eliminación",
            description=f"¿Estás seguro de que quieres eliminar el NPC **{npc['nombre']}**?",
            color=discord.Color.orange()
        )
        
        view = discord.ui.View()
        
        async def confirm_callback(button_interaction: discord.Interaction):
            await button_interaction.response.defer()
            
            # Eliminar mensaje del canal si existe
            if npc.get("mensaje_id") and npc.get("canal_id"):
                try:
                    channel = self.bot.get_channel(int(npc["canal_id"]))
                    if channel:
                        msg = await channel.fetch_message(int(npc["mensaje_id"]))
                        await msg.delete()
                except:
                    pass
            
            # Eliminar registro
            if self.bot.npc_manager.delete_record(npc_id):
                await button_interaction.followup.send(
                    f"✅ NPC **{npc['nombre']}** eliminado correctamente"
                )
            else:
                await button_interaction.followup.send("❌ Error al eliminar el NPC")
        
        async def cancel_callback(button_interaction: discord.Interaction):
            await button_interaction.response.edit_message(
                content="Eliminación cancelada",
                embed=None,
                view=None
            )
        
        confirm_btn = discord.ui.Button(label="Confirmar", style=discord.ButtonStyle.danger)
        cancel_btn = discord.ui.Button(label="Cancelar", style=discord.ButtonStyle.secondary)
        
        confirm_btn.callback = confirm_callback
        cancel_btn.callback = cancel_callback
        
        view.add_item(confirm_btn)
        view.add_item(cancel_btn)
        
        await interaction.followup.send(embed=embed, view=view)
    
    def crear_embed_npc(self, npc: Dict[str, Any], npc_id: int) -> discord.Embed:
        """Crear embed para mostrar información del NPC"""
        embed = discord.Embed(
            title=f"NPC: {npc['nombre']} (ID: {npc_id})",
            description=npc.get("descripcion", "Sin descripción"),
            color=discord.Color.blue()
        )
        
        if npc.get("imagen_url"):
            embed.set_image(url=npc["imagen_url"])
        
        embed.add_field(name="Rol", value=npc.get("rol", "N/A"), inline=True)
        
        # Inventario
        if npc.get("inventario"):
            item_ids = [int(x.strip()) for x in npc["inventario"].split(",") if x.strip()]
            item_names = []
            for item_id in item_ids:
                item = self.bot.item_manager.get_record_by_id(item_id)
                if item:
                    item_names.append(f"• {item['nombre']} (ID: {item_id})")
            
            inventario_text = "\n".join(item_names) if item_names else "Vacío"
            embed.add_field(name="Inventario", value=inventario_text[:1024], inline=False)
        else:
            embed.add_field(name="Inventario", value="Vacío", inline=False)
        
        embed.add_field(
            name="Canal asignado",
            value=f"<#{npc.get('canal_id')}>" if npc.get("canal_id") else "Sin canal",
            inline=True
        )
        
        embed.add_field(
            name="Creador",
            value=f"<@{npc.get('creador_id')}>" if npc.get("creador_id") else "Desconocido",
            inline=True
        )
        
        if npc.get("fecha_creacion"):
            try:
                fecha = datetime.fromisoformat(npc["fecha_creacion"])
                embed.set_footer(text=f"Creado el {fecha.strftime('%d/%m/%Y a las %H:%M')}")
            except:
                pass
        
        return embed
    
    @app_commands.command(name="npc_item_agregar", description="Agregar ítem a un NPC")
    @app_commands.describe(
        npc_id="ID del NPC",
        item_id="ID del ítem a agregar"
    )
    async def npc_item_agregar(self, interaction: discord.Interaction, npc_id: int, item_id: int):
        """Agregar ítem al inventario de un NPC"""
        await interaction.response.defer(ephemeral=True)
        
        # Verificar que existen ambos
        npc = self.bot.npc_manager.get_record_by_id(npc_id)
        item = self.bot.item_manager.get_record_by_id(item_id)
        
        if not npc:
            await interaction.followup.send("❌ NPC no encontrado")
            return
        
        if not item:
            await interaction.followup.send("❌ Ítem no encontrado")
            return
        
        # Verificar permisos
        if str(interaction.user.id) != npc.get("creador_id") and not interaction.user.guild_permissions.manage_channels:
            await interaction.followup.send("❌ No tienes permisos para modificar este NPC")
            return
        
        # Obtener inventario actual
        inventario_actual = npc.get("inventario", "")
        items_actuales = [int(x.strip()) for x in inventario_actual.split(",") if x.strip()]
        
        # Verificar si ya tiene el ítem
        if item_id in items_actuales:
            await interaction.followup.send(f"❌ El NPC ya tiene el ítem **{item['nombre']}**")
            return
        
        # Agregar ítem
        items_actuales.append(item_id)
        nuevo_inventario = ",".join(str(x) for x in items_actuales)
        
        # Actualizar
        if self.bot.npc_manager.update_record(npc_id, {"inventario": nuevo_inventario}):
            embed = discord.Embed(
                title="✅ Ítem Agregado",
                description=f"Se agregó **{item['nombre']}** al inventario de **{npc['nombre']}**",
                color=discord.Color.green()
            )
            await interaction.followup.send(embed=embed)
            
            # Actualizar ficha en el canal
            await self._actualizar_ficha_npc(npc_id)
        else:
            await interaction.followup.send("❌ Error al agregar el ítem")
    
    @app_commands.command(name="npc_item_quitar", description="Quitar ítem de un NPC")
    @app_commands.describe(
        npc_id="ID del NPC",
        item_id="ID del ítem a quitar"
    )
    async def npc_item_quitar(self, interaction: discord.Interaction, npc_id: int, item_id: int):
        """Quitar ítem del inventario de un NPC"""
        await interaction.response.defer(ephemeral=True)
        
        npc = self.bot.npc_manager.get_record_by_id(npc_id)
        if not npc:
            await interaction.followup.send("❌ NPC no encontrado")
            return
        
        # Verificar permisos
        if str(interaction.user.id) != npc.get("creador_id") and not interaction.user.guild_permissions.manage_channels:
            await interaction.followup.send("❌ No tienes permisos para modificar este NPC")
            return
        
        # Obtener inventario actual
        inventario_actual = npc.get("inventario", "")
        items_actuales = [int(x.strip()) for x in inventario_actual.split(",") if x.strip()]
        
        # Verificar si tiene el ítem
        if item_id not in items_actuales:
            await interaction.followup.send("❌ El NPC no tiene ese ítem")
            return
        
        # Quitar ítem
        items_actuales.remove(item_id)
        nuevo_inventario = ",".join(str(x) for x in items_actuales)
        
        # Obtener nombre del ítem
        item = self.bot.item_manager.get_record_by_id(item_id)
        item_nombre = item['nombre'] if item else f"ID: {item_id}"
        
        # Actualizar
        if self.bot.npc_manager.update_record(npc_id, {"inventario": nuevo_inventario}):
            embed = discord.Embed(
                title="✅ Ítem Quitado",
                description=f"Se quitó **{item_nombre}** del inventario de **{npc['nombre']}**",
                color=discord.Color.green()
            )
            await interaction.followup.send(embed=embed)
            
            # Actualizar ficha en el canal
            await self._actualizar_ficha_npc(npc_id)
        else:
            await interaction.followup.send("❌ Error al quitar el ítem")
    
    async def _actualizar_ficha_npc(self, npc_id: int):
        """Actualizar la ficha del NPC en su canal asignado"""
        npc = self.bot.npc_manager.get_record_by_id(npc_id)
        if npc and npc.get("mensaje_id") and npc.get("canal_id"):
            try:
                channel = self.bot.get_channel(int(npc["canal_id"]))
                if channel:
                    msg = await channel.fetch_message(int(npc["mensaje_id"]))
                    new_embed = self.crear_embed_npc(npc, npc_id)
                    await msg.edit(embed=new_embed)
            except Exception as e:
                logger.error(f"Error al actualizar ficha del NPC {npc_id}: {e}")

class ItemCommands(commands.Cog):
    def __init__(self, bot: NPCBot):
        self.bot = bot
    
    @app_commands.command(name="item_crear", description="Crea un nuevo ítem")
    async def item_crear(self, interaction: discord.Interaction):
        """Crear un nuevo ítem"""
        modal = ItemModal(title="Crear Nuevo Ítem")
        
        async def on_submit(modal_interaction: discord.Interaction):
            await modal_interaction.response.defer(ephemeral=True)
            
            # Crear vista para seleccionar categoría
            view = discord.ui.View()
            categoria_select = CategoriaSelect()
            view.add_item(categoria_select)
            
            embed = discord.Embed(
                title="Configuración del Ítem",
                description="Selecciona la categoría del ítem",
                color=discord.Color.green()
            )
            
            await modal_interaction.followup.send(embed=embed, view=view, ephemeral=True)
            
            # Esperar selección
            def check(i):
                return i.user == interaction.user
            
            try:
                select_interaction = await self.bot.wait_for("interaction", check=check, timeout=60)
                await select_interaction.response.defer()
                
                categoria = categoria_select.values[0] if categoria_select.values else "Objeto"
                
                # Crear ítem
                item_data = {
                    "nombre": modal.nombre.value,
                    "descripcion": modal.descripcion.value,
                    "imagen_url": modal.imagen_url.value or "",
                    "categoria": categoria,
                    "propiedades": modal.propiedades.value or "",
                    "creador_id": str(interaction.user.id)
                }
                
                item_id = self.bot.item_manager.add_record(item_data)
                
                confirm_embed = discord.Embed(
                    title="✅ Ítem Creado Exitosamente",
                    description=f"**{item_data['nombre']}** (ID: {item_id})",
                    color=discord.Color.green()
                )
                confirm_embed.add_field(name="Categoría", value=categoria, inline=True)
                if item_data["imagen_url"]:
                    confirm_embed.set_thumbnail(url=item_data["imagen_url"])
                
                await select_interaction.followup.send(embed=confirm_embed)
                
            except asyncio.TimeoutError:
                await modal_interaction.followup.send("⏱️ Tiempo agotado", ephemeral=True)
        
        modal.on_submit = on_submit
        await interaction.response.send_modal(modal)
    
    @app_commands.command(name="item_listar", description="Lista todos los ítems")
    @app_commands.describe(categoria="Filtrar por categoría (opcional)")
    async def item_listar(self, interaction: discord.Interaction,
                         categoria: Optional[str] = None):
        """Listar ítems con paginación"""
        await interaction.response.defer()
        
        items = self.bot.item_manager.get_all_records()
        
        if categoria:
            items = [item for item in items if item.get("categoria") == categoria]
        
        if not items:
            await interaction.followup.send("❌ No se encontraron ítems")
            return
        
        # Crear embeds paginados
        embeds = []
        items_per_page = 5
        
        for i in range(0, len(items), items_per_page):
            embed = discord.Embed(
                title=f"📦 Lista de Ítems" + (f" - Categoría: {categoria}" if categoria else ""),
                color=discord.Color.green()
            )
            
            page_items = items[i:i + items_per_page]
            for item in page_items:
                creador_text = f"<@{item.get('creador_id')}>" if item.get('creador_id') else "Desconocido"
                
                embed.add_field(
                    name=f"{item['nombre']} (ID: {item['id']})",
                    value=f"**Categoría:** {item.get('categoria', 'N/A')}\n"
                          f"**Creador:** {creador_text}",
                    inline=False
                )
            
            embed.set_footer(text=f"Página {len(embeds) + 1} de {((len(items) - 1) // items_per_page) + 1}")
            embeds.append(embed)
        
        if len(embeds) == 1:
            await interaction.followup.send(embed=embeds[0])
        else:
            view = PaginationView(embeds)
            await interaction.followup.send(embed=embeds[0], view=view)
    
    @app_commands.command(name="item_ver", description="Ver detalles de un ítem")
    @app_commands.describe(item_id="ID del ítem")
    async def item_ver(self, interaction: discord.Interaction, item_id: int):
        """Ver información detallada de un ítem"""
        await interaction.response.defer()
        
        item = self.bot.item_manager.get_record_by_id(item_id)
        if not item:
            await interaction.followup.send("❌ Ítem no encontrado")
            return
        
        embed = self.crear_embed_item(item, item_id)
        await interaction.followup.send(embed=embed)
    
    @app_commands.command(name="item_editar", description="Editar un ítem existente")
    @app_commands.describe(item_id="ID del ítem a editar")
    async def item_editar(self, interaction: discord.Interaction, item_id: int):
        """Editar un ítem"""
        item = self.bot.item_manager.get_record_by_id(item_id)
        if not item:
            await interaction.response.send_message("❌ Ítem no encontrado", ephemeral=True)
            return
        
        # Verificar permisos
        if str(interaction.user.id) != item.get("creador_id") and not interaction.user.guild_permissions.manage_channels:
            await interaction.response.send_message("❌ No tienes permisos para editar este ítem", ephemeral=True)
            return
        
        modal = ItemModal(title=f"Editar Ítem: {item['nombre']}", item_data=item)
        
        async def on_submit(modal_interaction: discord.Interaction):
            await modal_interaction.response.defer(ephemeral=True)
            
            updates = {
                "nombre": modal.nombre.value,
                "descripcion": modal.descripcion.value,
                "imagen_url": modal.imagen_url.value or "",
                "propiedades": modal.propiedades.value or ""
            }
            
            if self.bot.item_manager.update_record(item_id, updates):
                embed = discord.Embed(
                    title="✅ Ítem Actualizado",
                    description=f"**{updates['nombre']}** ha sido actualizado",
                    color=discord.Color.green()
                )
                await modal_interaction.followup.send(embed=embed)
            else:
                await modal_interaction.followup.send("❌ Error al actualizar el ítem", ephemeral=True)
        
        modal.on_submit = on_submit
        await interaction.response.send_modal(modal)
    
    @app_commands.command(name="item_eliminar", description="Eliminar un ítem")
    @app_commands.describe(item_id="ID del ítem a eliminar")
    async def item_eliminar(self, interaction: discord.Interaction, item_id: int):
        """Eliminar un ítem"""
        await interaction.response.defer(ephemeral=True)
        
        item = self.bot.item_manager.get_record_by_id(item_id)
        if not item:
            await interaction.followup.send("❌ Ítem no encontrado")
            return
        
        # Verificar permisos
        if str(interaction.user.id) != item.get("creador_id") and not interaction.user.guild_permissions.manage_channels:
            await interaction.followup.send("❌ No tienes permisos para eliminar este ítem")
            return
        
        # Confirmación
        embed = discord.Embed(
            title="⚠️ Confirmar Eliminación",
            description=f"¿Estás seguro de que quieres eliminar el ítem **{item['nombre']}**?",
            color=discord.Color.orange()
        )
        
        view = discord.ui.View()
        
        async def confirm_callback(button_interaction: discord.Interaction):
            await button_interaction.response.defer()
            
            if self.bot.item_manager.delete_record(item_id):
                await button_interaction.followup.send(
                    f"✅ Ítem **{item['nombre']}** eliminado correctamente"
                )
            else:
                await button_interaction.followup.send("❌ Error al eliminar el ítem")
        
        async def cancel_callback(button_interaction: discord.Interaction):
            await button_interaction.response.edit_message(
                content="Eliminación cancelada",
                embed=None,
                view=None
            )
        
        confirm_btn = discord.ui.Button(label="Confirmar", style=discord.ButtonStyle.danger)
        cancel_btn = discord.ui.Button(label="Cancelar", style=discord.ButtonStyle.secondary)
        
        confirm_btn.callback = confirm_callback
        cancel_btn.callback = cancel_callback
        
        view.add_item(confirm_btn)
        view.add_item(cancel_btn)
        
        await interaction.followup.send(embed=embed, view=view)
    
    def crear_embed_item(self, item: Dict[str, Any], item_id: int) -> discord.Embed:
        """Crear embed para mostrar información del ítem"""
        embed = discord.Embed(
            title=f"Ítem: {item['nombre']} (ID: {item_id})",
            description=item.get("descripcion", "Sin descripción"),
            color=discord.Color.green()
        )
        
        if item.get("imagen_url"):
            embed.set_thumbnail(url=item["imagen_url"])
        
        embed.add_field(name="Categoría", value=item.get("categoria", "N/A"), inline=True)
        
        # Propiedades
        if item.get("propiedades"):
            props_list = []
            for prop in item["propiedades"].split(";"):
                if ":" in prop:
                    key, value = prop.split(":", 1)
                    props_list.append(f"• **{key.strip()}:** {value.strip()}")
            
            props_text = "\n".join(props_list) if props_list else "Sin propiedades"
            embed.add_field(name="Propiedades", value=props_text[:1024], inline=False)
        
        embed.add_field(
            name="Creador",
            value=f"<@{item.get('creador_id')}>" if item.get("creador_id") else "Desconocido",
            inline=True
        )
        
        if item.get("fecha_creacion"):
            try:
                fecha = datetime.fromisoformat(item["fecha_creacion"])
                embed.set_footer(text=f"Creado el {fecha.strftime('%d/%m/%Y a las %H:%M')}")
            except:
                pass
        
        return embed

# Inicializar y ejecutar el bot
bot = NPCBot()

@bot.event
async def on_ready():
    logger.info(f"Bot conectado como {bot.user} (ID: {bot.user.id})")
    try:
        synced = await bot.tree.sync()
        logger.info(f"Sincronizados {len(synced)} comandos slash")
    except Exception as e:
        logger.error(f"Error al sincronizar comandos: {e}")

if __name__ == "__main__":
    if not DISCORD_TOKEN:
        logger.error("No se encontró DISCORD_TOKEN en .env")
    else:
        bot.run(DISCORD_TOKEN)